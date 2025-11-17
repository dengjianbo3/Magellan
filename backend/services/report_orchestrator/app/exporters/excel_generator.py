"""
Excel Report Generator
Excel报告生成器

Uses openpyxl to generate structured investment analysis reports in Excel format.
使用openpyxl生成结构化的投资分析报告Excel格式。
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, Any, List


class ExcelReportGenerator:
    """
    Excel报告生成器

    Features:
    - 多工作表结构
    - 中英文支持
    - 数据表格化
    - 专业样式
    - 易于数据分析
    """

    def __init__(self, language: str = "zh"):
        """
        初始化Excel生成器

        Args:
            language: 语言 ("zh" 中文, "en" 英文)
        """
        self.language = language
        self.wb = Workbook()
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # Define styles
        self.header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        self.subheader_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.subheader_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
        self.title_font = Font(name='Arial', size=11, bold=True)
        self.normal_font = Font(name='Arial', size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_report(
        self,
        report_data: Dict[str, Any],
        output_path: str
    ) -> str:
        """
        生成完整Excel报告

        Args:
            report_data: 报告数据
            output_path: 输出文件路径

        Returns:
            生成的Excel文件路径
        """
        # 1. Summary Sheet
        self._add_summary_sheet(report_data)

        # 2. Financial Data Sheet
        if report_data.get('preliminary_im', {}).get('financial_highlights'):
            self._add_financial_sheet(report_data['preliminary_im'])

        # 3. Market Analysis Sheet
        if report_data.get('market_analysis'):
            self._add_market_sheet(report_data['market_analysis'])

        # 4. Team Analysis Sheet
        if report_data.get('team_analysis'):
            self._add_team_sheet(report_data['team_analysis'])

        # 5. Risk Assessment Sheet
        if report_data.get('preliminary_im', {}).get('risks'):
            self._add_risk_sheet(report_data['preliminary_im'])

        # 6. Raw Data Sheet (full JSON dump for reference)
        self._add_raw_data_sheet(report_data)

        # Save workbook
        self.wb.save(output_path)
        return output_path

    def _add_summary_sheet(self, report_data: Dict[str, Any]):
        """添加摘要工作表"""
        if self.language == "en":
            ws = self.wb.create_sheet("Executive Summary")
        else:
            ws = self.wb.create_sheet("执行摘要")

        row = 1

        # Title
        if self.language == "en":
            ws.cell(row, 1, "Investment Analysis Report - Executive Summary")
        else:
            ws.cell(row, 1, "投资尽职调查报告 - 执行摘要")
        ws.cell(row, 1).font = Font(size=16, bold=True)
        row += 2

        # Company Info
        company_name = report_data.get('company_name', 'N/A')
        if self.language == "en":
            ws.cell(row, 1, "Company:")
        else:
            ws.cell(row, 1, "公司:")
        ws.cell(row, 1).font = self.title_font
        ws.cell(row, 2, company_name)
        row += 1

        # Date
        date_str = datetime.now().strftime("%Y-%m-%d")
        if self.language == "en":
            ws.cell(row, 1, "Report Date:")
        else:
            ws.cell(row, 1, "报告日期:")
        ws.cell(row, 1).font = self.title_font
        ws.cell(row, 2, date_str)
        row += 2

        # Investment Recommendation
        im_data = report_data.get('preliminary_im', {})
        if im_data.get('investment_recommendation'):
            if self.language == "en":
                ws.cell(row, 1, "Investment Recommendation:")
            else:
                ws.cell(row, 1, "投资建议:")
            ws.cell(row, 1).font = self.subheader_font
            ws.cell(row, 1).fill = self.subheader_fill
            ws.merge_cells(f'A{row}:E{row}')
            row += 1

            ws.cell(row, 1, str(im_data['investment_recommendation']))
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical='top')
            row += 2

        # Key Findings
        if im_data.get('key_findings'):
            if self.language == "en":
                ws.cell(row, 1, "Key Findings:")
            else:
                ws.cell(row, 1, "关键发现:")
            ws.cell(row, 1).font = self.subheader_font
            ws.cell(row, 1).fill = self.subheader_fill
            ws.merge_cells(f'A{row}:E{row}')
            row += 1

            findings = im_data['key_findings']
            if isinstance(findings, list):
                for finding in findings:
                    ws.cell(row, 1, "•")
                    ws.cell(row, 2, str(finding))
                    ws.merge_cells(f'B{row}:E{row}')
                    ws.cell(row, 2).alignment = Alignment(wrap_text=True)
                    row += 1
            else:
                ws.cell(row, 1, str(findings))
                ws.merge_cells(f'A{row}:E{row}')
                ws.cell(row, 1).alignment = Alignment(wrap_text=True)
                row += 1
            row += 1

        # Investment Highlights
        if im_data.get('investment_highlights'):
            if self.language == "en":
                ws.cell(row, 1, "Investment Highlights:")
            else:
                ws.cell(row, 1, "投资亮点:")
            ws.cell(row, 1).font = self.subheader_font
            ws.cell(row, 1).fill = self.subheader_fill
            ws.merge_cells(f'A{row}:E{row}')
            row += 1

            highlights = im_data['investment_highlights']
            if isinstance(highlights, list):
                for hl in highlights:
                    ws.cell(row, 1, "•")
                    ws.cell(row, 2, str(hl))
                    ws.merge_cells(f'B{row}:E{row}')
                    ws.cell(row, 2).alignment = Alignment(wrap_text=True)
                    row += 1
            else:
                ws.cell(row, 1, str(highlights))
                ws.merge_cells(f'A{row}:E{row}')
                ws.cell(row, 1).alignment = Alignment(wrap_text=True)
                row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

    def _add_financial_sheet(self, im_data: Dict[str, Any]):
        """添加财务数据工作表"""
        if self.language == "en":
            ws = self.wb.create_sheet("Financial Data")
        else:
            ws = self.wb.create_sheet("财务数据")

        row = 1

        # Title
        if self.language == "en":
            ws.cell(row, 1, "Financial Analysis")
        else:
            ws.cell(row, 1, "财务分析")
        ws.cell(row, 1).font = Font(size=14, bold=True)
        row += 2

        # Financial Highlights Table
        fin_data = im_data.get('financial_highlights', {})
        if fin_data:
            if self.language == "en":
                ws.cell(row, 1, "Metric")
                ws.cell(row, 2, "Value")
            else:
                ws.cell(row, 1, "指标")
                ws.cell(row, 2, "数值")

            ws.cell(row, 1).font = self.subheader_font
            ws.cell(row, 1).fill = self.subheader_fill
            ws.cell(row, 2).font = self.subheader_font
            ws.cell(row, 2).fill = self.subheader_fill
            row += 1

            for key, value in fin_data.items():
                ws.cell(row, 1, str(key))
                ws.cell(row, 2, str(value))
                ws.cell(row, 1).border = self.border
                ws.cell(row, 2).border = self.border
                row += 1

            row += 1

        # Financial Analysis Text
        if im_data.get('financial_analysis'):
            if self.language == "en":
                ws.cell(row, 1, "Analysis:")
            else:
                ws.cell(row, 1, "分析:")
            ws.cell(row, 1).font = self.title_font
            row += 1

            ws.cell(row, 1, str(im_data['financial_analysis']))
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical='top')

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _add_market_sheet(self, market_data: Dict[str, Any]):
        """添加市场分析工作表"""
        if self.language == "en":
            ws = self.wb.create_sheet("Market Analysis")
        else:
            ws = self.wb.create_sheet("市场分析")

        row = 1

        # Title
        if self.language == "en":
            ws.cell(row, 1, "Market Analysis")
        else:
            ws.cell(row, 1, "市场分析")
        ws.cell(row, 1).font = Font(size=14, bold=True)
        row += 2

        # Market Size
        if market_data.get('market_size_analysis'):
            if self.language == "en":
                ws.cell(row, 1, "Market Size:")
            else:
                ws.cell(row, 1, "市场规模:")
            ws.cell(row, 1).font = self.title_font
            row += 1

            ws.cell(row, 1, str(market_data['market_size_analysis']))
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical='top')
            row += 2

        # Competitive Landscape
        if market_data.get('competitive_landscape'):
            if self.language == "en":
                ws.cell(row, 1, "Competitive Landscape:")
            else:
                ws.cell(row, 1, "竞争格局:")
            ws.cell(row, 1).font = self.title_font
            row += 1

            ws.cell(row, 1, str(market_data['competitive_landscape']))
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical='top')
            row += 2

        # Market Trends
        if market_data.get('market_trends'):
            if self.language == "en":
                ws.cell(row, 1, "Market Trends:")
            else:
                ws.cell(row, 1, "市场趋势:")
            ws.cell(row, 1).font = self.title_font
            row += 1

            trends = market_data['market_trends']
            if isinstance(trends, list):
                for trend in trends:
                    ws.cell(row, 1, "•")
                    ws.cell(row, 2, str(trend))
                    ws.merge_cells(f'B{row}:D{row}')
                    ws.cell(row, 2).alignment = Alignment(wrap_text=True)
                    row += 1
            else:
                ws.cell(row, 1, str(trends))
                ws.merge_cells(f'A{row}:D{row}')
                ws.cell(row, 1).alignment = Alignment(wrap_text=True)
                row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _add_team_sheet(self, team_data: Dict[str, Any]):
        """添加团队分析工作表"""
        if self.language == "en":
            ws = self.wb.create_sheet("Team Analysis")
        else:
            ws = self.wb.create_sheet("团队分析")

        row = 1

        # Title
        if self.language == "en":
            ws.cell(row, 1, "Team Assessment")
        else:
            ws.cell(row, 1, "团队评估")
        ws.cell(row, 1).font = Font(size=14, bold=True)
        row += 2

        # Team Score
        if team_data.get('experience_match_score'):
            if self.language == "en":
                ws.cell(row, 1, "Experience Match Score:")
            else:
                ws.cell(row, 1, "经验匹配度评分:")
            ws.cell(row, 1).font = self.title_font
            ws.cell(row, 2, f"{team_data['experience_match_score']}/10")
            ws.cell(row, 2).font = Font(size=12, bold=True, color='E74C3C')
            row += 2

        # Team Strengths
        if team_data.get('strengths'):
            if self.language == "en":
                ws.cell(row, 1, "Team Strengths:")
            else:
                ws.cell(row, 1, "团队优势:")
            ws.cell(row, 1).font = self.title_font
            row += 1

            strengths = team_data['strengths']
            if isinstance(strengths, list):
                for strength in strengths:
                    ws.cell(row, 1, "•")
                    ws.cell(row, 2, str(strength))
                    ws.merge_cells(f'B{row}:D{row}')
                    ws.cell(row, 2).alignment = Alignment(wrap_text=True)
                    row += 1
            else:
                ws.cell(row, 1, str(strengths))
                ws.merge_cells(f'A{row}:D{row}')
                ws.cell(row, 1).alignment = Alignment(wrap_text=True)
                row += 1
            row += 1

        # Recommendations
        if team_data.get('recommendations'):
            if self.language == "en":
                ws.cell(row, 1, "Recommendations:")
            else:
                ws.cell(row, 1, "改进建议:")
            ws.cell(row, 1).font = self.title_font
            row += 1

            recs = team_data['recommendations']
            if isinstance(recs, list):
                for rec in recs:
                    ws.cell(row, 1, "•")
                    ws.cell(row, 2, str(rec))
                    ws.merge_cells(f'B{row}:D{row}')
                    ws.cell(row, 2).alignment = Alignment(wrap_text=True)
                    row += 1
            else:
                ws.cell(row, 1, str(recs))
                ws.merge_cells(f'A{row}:D{row}')
                ws.cell(row, 1).alignment = Alignment(wrap_text=True)
                row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _add_risk_sheet(self, im_data: Dict[str, Any]):
        """添加风险评估工作表"""
        if self.language == "en":
            ws = self.wb.create_sheet("Risk Assessment")
        else:
            ws = self.wb.create_sheet("风险评估")

        row = 1

        # Title
        if self.language == "en":
            ws.cell(row, 1, "Risk Evaluation")
        else:
            ws.cell(row, 1, "风险评估")
        ws.cell(row, 1).font = Font(size=14, bold=True)
        row += 2

        # Risks Table
        risks = im_data.get('risks', [])
        if isinstance(risks, list) and risks:
            # Table header
            if self.language == "en":
                ws.cell(row, 1, "Risk Name")
                ws.cell(row, 2, "Level")
                ws.cell(row, 3, "Description")
            else:
                ws.cell(row, 1, "风险名称")
                ws.cell(row, 2, "等级")
                ws.cell(row, 3, "描述")

            for col in range(1, 4):
                ws.cell(row, col).font = self.subheader_font
                ws.cell(row, col).fill = self.subheader_fill
            row += 1

            for risk in risks:
                if isinstance(risk, dict):
                    risk_name = risk.get('name', 'Unknown Risk')
                    risk_level = risk.get('level', 'Unknown')
                    risk_desc = risk.get('description', '')

                    ws.cell(row, 1, risk_name)
                    ws.cell(row, 2, risk_level)
                    ws.cell(row, 3, risk_desc)

                    for col in range(1, 4):
                        ws.cell(row, col).border = self.border
                        ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical='top')

                    row += 1
                else:
                    ws.cell(row, 1, "•")
                    ws.cell(row, 2, str(risk))
                    ws.merge_cells(f'B{row}:C{row}')
                    ws.cell(row, 2).alignment = Alignment(wrap_text=True)
                    row += 1
        else:
            ws.cell(row, 1, str(risks))
            ws.merge_cells(f'A{row}:C{row}')
            ws.cell(row, 1).alignment = Alignment(wrap_text=True)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 60

    def _add_raw_data_sheet(self, report_data: Dict[str, Any]):
        """添加原始数据工作表（供参考）"""
        if self.language == "en":
            ws = self.wb.create_sheet("Raw Data")
        else:
            ws = self.wb.create_sheet("原始数据")

        row = 1

        # Title
        if self.language == "en":
            ws.cell(row, 1, "Raw Report Data (JSON)")
        else:
            ws.cell(row, 1, "原始报告数据 (JSON)")
        ws.cell(row, 1).font = Font(size=14, bold=True)
        row += 2

        # Add JSON data as text
        import json
        json_str = json.dumps(report_data, indent=2, ensure_ascii=False)

        ws.cell(row, 1, json_str)
        ws.merge_cells(f'A{row}:E{row}')
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row, 1).font = Font(name='Courier New', size=9)

        ws.column_dimensions['A'].width = 100


# 便捷函数
def generate_excel_report(
    report_data: Dict[str, Any],
    output_path: str,
    language: str = "zh"
) -> str:
    """
    生成Excel报告的便捷函数

    Args:
        report_data: 报告数据
        output_path: 输出文件路径
        language: 语言设置

    Returns:
        生成的Excel文件路径
    """
    generator = ExcelReportGenerator(language=language)
    return generator.generate_report(report_data, output_path)
